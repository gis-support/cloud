from tempfile import NamedTemporaryFile

from app.helpers.layer import Layer
from app.helpers.cloud import Cloud

import openpyxl

def get_xlsx(layer: Layer, feature_id: int, buffer_distance: float, feature_name: str) -> NamedTemporaryFile:
    distances = layer.get_distances(feature_id, buffer_distance)

    title_rows = [
        ["Nazwa warstwy: ", layer.name],
        ["Nazwa obiektu: ", feature_name]
    ]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 25

    for title in title_rows:
        sheet.append(title)

    sheet.append([])

    distances = distances["pn"] + distances["pk"]
    distances_headers = ["Forma ochrony przyrody", "Odległość [km]"]
    sheet.append(distances_headers)
    for dist in distances:
        object_name = dist["name"]
        distance = dist["distance"]

        distance_km = distance/1000

        sheet.append([object_name, round(distance_km, 2)])

    result = NamedTemporaryFile()
    workbook.save(result.name)

    for row_number in range(1, sheet.max_row):
        sheet.row_dimensions[row_number] = 30

    return result


def get_xlsx_geojson(cloud: Cloud, geometry: str, buffer_distance: float, teryt: str) -> NamedTemporaryFile:
    distances = cloud.get_distances(geometry=geometry, buffer=buffer_distance)

    title_rows = [
        ["Nazwa warstwy: ", "ULDK"],
        ["Nazwa obiektu: ", teryt]
    ]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 25

    for title in title_rows:
        sheet.append(title)

    sheet.append([])

    distances = distances["pn"] + distances["pk"]
    distances_headers = ["Forma ochrony przyrody", "Odległość [km]"]
    sheet.append(distances_headers)
    for dist in distances:
        object_name = dist["name"]
        distance = dist["distance"]

        distance_km = distance/1000

        sheet.append([object_name, round(distance_km, 2)])

    result = NamedTemporaryFile()
    workbook.save(result.name)

    for row_number in range(1, sheet.max_row):
        sheet.row_dimensions[row_number] = 30

    return result